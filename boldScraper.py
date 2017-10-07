# Import libraries
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Load the WorkBook
wb = load_workbook(filename='TAXDB_WebCrawling.xlsx', data_only=True)
# Create two sheets to separate the links into
wb.create_sheet(title="valid_links")
wb.create_sheet(title="invalid_links")
# Select the Sheets to work with
sheet = wb['Sheet1']
valid_links = wb['valid_links']
invalid_links = wb['invalid_links']
# Create 2 lists that store links that contain content and links with no content
valid = []
invalid = []
# Start working on the cells
for cell in range(2, sheet.max_row+1):
    if sheet.cell(row=cell, column=4).value is not None:
        e = sheet.cell(row=cell, column=4)
        # specify the URL to the main BOLD page
        original_page = e.value
        print(original_page)
        # Select name of the taxon
        name = sheet.cell(row=cell, column=3).value
        # Use the BOLD JS call to the Wikipedia API for the data
        # And use the rank in the xlsx file to form the API call
        rank = sheet.cell(row=cell, column=2).value
        # Convert the rank from an int to a string
        rank = str(rank)
        # Create the Link
        URL = "http://boldsystems.org/index.php/MAS_Ajax_WikiRetriever?name=" + name + "&rank=" + rank
        # Again, query the website and return the html to the variable ‘page_check’
        page_check = urllib.request.urlopen(URL)
        # parse the html using beautiful soup and store in variable `soup2`
        soup2 = BeautifulSoup(page_check, 'html.parser')
        # simply filter out the links to different lists we created.
        if len(soup2) <= 1 or soup2 is "null":
            invalid.append(original_page)
        else:
            valid.append(original_page)

for i in range(0, len(valid)):
    e = valid_links.cell(row=1+i, column=1)
    e.value = valid[i]
for k in range(0, len(invalid)):
    o = invalid_links.cell(row=1+k, column=1)
    o.value = invalid[k]

wb.save("TAXDB_WebCrawling.xlsx")
print("Done")
